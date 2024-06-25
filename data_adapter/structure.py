from __future__ import annotations

import re
from typing import List, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from data_adapter import settings

MAX_IDENTIFIER_LENGTH = 50

# pylint:disable=C0209
IDENTIFIER_PATTERN = re.compile("^[a-z][a-z0-9_, ]{0,%s}$" % (MAX_IDENTIFIER_LENGTH - 1))


class StructureError(Exception):
    """Raised if structure is corrupted."""


def check_character_convention(dataframe: pd.DataFrame, cols: Optional[List[str]] = None):
    """Check in parameter-, process-, input-and output-column for character convention.

    Parameters
    ----------
    dataframe: pandas.DataFrame
        Parameters in dataframe are checked for convention
    cols: Optional[List[str]]
        Columns to check for

    Raises
    ------
    ValueError
        if element in dataframe does not fit character convention

    """
    cols = dataframe.columns if cols is None else cols
    for col in cols:
        for element in dataframe[col]:
            if not isinstance(element, str):
                continue
            if not IDENTIFIER_PATTERN.match(element):
                raise ValueError(f"Wrong syntax: {element}\nAllowed are characters: a-z and 0-9 and , and _")


def _initialize_commodities(sectors):
    return {sector: set() for sector in sectors}


def _process_row(
    process_name, io_dict, category, input_commodities, output_commodities, source_commodities, import_commodities
):
    row_sectors = [process_name.split("_")[0]]
    inputs = [num for item in io_dict["inputs"] for num in (item if isinstance(item, list) else (item,))]
    outputs = [num for item in io_dict["outputs"] for num in (item if isinstance(item, list) else (item,))]

    categories_set = set([category.split("_")[0] for category in inputs + outputs])

    sources = [
        num
        for item in io_dict["outputs"]
        if "source" in process_name
        for num in (item if isinstance(item, list) else (item,))
    ]
    imports = [
        num
        for item in io_dict["outputs"]
        if "import" in process_name
        for num in (item if isinstance(item, list) else (item,))
    ]

    for sector in row_sectors:
        if category in categories_set:
            input_commodities[sector] |= set([commodity for commodity in inputs if commodity.startswith(category)])
            source_commodities[sector] |= set([commodity for commodity in sources if commodity.startswith(category)])
            import_commodities[sector] |= set([commodity for commodity in imports if commodity.startswith(category)])
            output_commodities[sector] |= set([commodity for commodity in outputs if commodity.startswith(category)])


def _create_matrix_data(input_commodities, output_commodities, source_commodities, import_commodities, sectors):
    combined = {
        sector: list(
            input_commodities[sector]
            | output_commodities[sector]
            | source_commodities[sector]
            | import_commodities[sector]
        )
        for sector in sectors
    }

    unique_commodity_list = sorted(list(set(commodity for sector in combined.values() for commodity in sector)))

    matrix_data = pd.DataFrame(np.NaN, index=unique_commodity_list, columns=sectors)

    # Populate matrix_data based on conditions
    for sector in sectors:
        for commodity in combined[sector]:
            # if commodity in import_commodities[sector]:
            #     matrix_data.at[commodity, sector] = 3
            if commodity in input_commodities[sector] and commodity in output_commodities[sector]:
                matrix_data.at[commodity, sector] = 0
            elif commodity in input_commodities[sector]:
                matrix_data.at[commodity, sector] = -1
            elif commodity in output_commodities[sector]:
                matrix_data.at[commodity, sector] = 1
            if commodity in source_commodities[sector] and commodity in output_commodities[sector]:
                matrix_data.at[commodity, sector] = 3
            elif commodity in source_commodities[sector]:
                matrix_data.at[commodity, sector] = 2

    return matrix_data


def _process_data(category, sectors, processes):
    input_commodities = _initialize_commodities(sectors)
    output_commodities = _initialize_commodities(sectors)
    source_commodities = _initialize_commodities(sectors)
    import_commodities = _initialize_commodities(sectors)

    for process_name, io_dict in processes.items():
        if any([keyword in process_name for keyword in ["storage", "export", "delivery"]]):
            continue

        _process_row(
            process_name,
            io_dict,
            category,
            input_commodities,
            output_commodities,
            source_commodities,
            import_commodities,
        )

    matrix_data = _create_matrix_data(
        input_commodities, output_commodities, source_commodities, import_commodities, sectors
    )

    return matrix_data


class Structure:
    def __init__(
        self,
        structure_name: str,
        process_sheet: str = "Process_Set",
        parameter_sheet: str = "Parameter_Input-Output",
        helper_sheet: str = "Helper_Set",
    ):
        self.structure_file = settings.STRUCTURES_DIR / f"{structure_name}.xlsx"
        self.processes = self._init_processes(process_sheet, helper_sheet)
        self.parameters = self._init_parameters(parameter_sheet)

    def _init_processes(self, process_sheet: str, helper_sheet: str) -> dict:
        """Parse (helper) processes with corresponding inputs and outputs to dict.

        Parameters
        ----------
        process_sheet: str
            Sheet to read processes from
        helper_sheet: str
            Sheet to read additional helper processes from

        Returns
        -------
        dict
            Energy modelling processes and related inputs and outputs
        """

        def get_nodes(nodes_raw):
            nodes_raw_stripped = nodes_raw.replace(" ", "")
            grouped_nodes_raw = re.findall(r"\[[\w*,\,]*]", nodes_raw_stripped)

            nodes = []
            for group in grouped_nodes_raw:
                nodes.append(get_nodes(group[1:-1]))  # without brackets
                # Remove groups
                nodes_raw_stripped = nodes_raw_stripped.replace(group, "")
            nodes += [node for node in nodes_raw_stripped.split(",") if node != ""]
            return nodes

        processes_raw = pd.read_excel(
            io=self.structure_file,
            sheet_name=process_sheet,
            usecols=("process", "input", "output"),
        )
        wb = load_workbook(self.structure_file, read_only=True)
        if helper_sheet in wb.sheetnames:
            helpers_raw = pd.read_excel(
                io=self.structure_file,
                sheet_name=helper_sheet,
                usecols=("process", "input", "output"),
            )
            processes_raw = pd.concat([processes_raw, helpers_raw])
        processes_raw = processes_raw.fillna("")
        check_character_convention(processes_raw, ["process"])
        processes = processes_raw.to_dict(orient="records")
        return {
            process["process"]: {"inputs": get_nodes(process["input"]), "outputs": get_nodes(process["output"])}
            for process in processes
        }

    def _init_parameters(self, parameter_sheet: str) -> dict:
        """Parse processes and its parameters with corresponding inputs and outputs to dict.

        Parameters
        ----------
        parameter_sheet: str
            Sheet to read parameters from

        Returns
        -------
        dict
            Energy modelling processes, its parameters and inputs and output
        """
        process_parameter_in_out = pd.read_excel(
            io=self.structure_file,
            sheet_name=parameter_sheet,
            usecols=("parameter", "process", "inputs", "outputs"),
        )
        process_parameter_in_out = process_parameter_in_out.fillna("")
        check_character_convention(process_parameter_in_out, ["process", "parameter"])

        # create ES_STRUCTURE dict from process_parameter_in_out
        list_dic = process_parameter_in_out.to_dict(orient="records")

        es_structure = {}

        for dic in list_dic:
            dic_para = {}

            if isinstance(dic.get("inputs"), str):
                inputs = {"inputs": dic.get("inputs").replace(" ", "").split(",")}
            else:
                inputs = {"inputs": []}
            if isinstance(dic.get("outputs"), str):
                outputs = {"outputs": dic.get("outputs").replace(" ", "").split(",")}
            else:
                outputs = {"outputs": []}

            dic_para[dic.get("parameter")] = inputs | outputs

            if dic.get("process") not in es_structure:
                es_structure[dic.get("process")] = dic_para
            else:
                es_structure[dic.get("process")] = es_structure[dic.get("process")] | dic_para

        return es_structure

    def plot_commodity_interfaces(
        self,
        categories=["pri", "sec", "iip", "exo", "emi"],
        sectors=["pow", "x2x", "ind", "mob", "hea", "helper"],
    ):
        try:
            import matplotlib.pyplot as plt
            from matplotlib.patches import Patch
        except ImportError:
            raise ImportError("You must install matplotlib in order to use this functionality.")

        cols = len(categories)
        fig, axes = plt.subplots(nrows=1, ncols=cols, figsize=(cols * 10, 16), sharey=False)
        if len(categories) == 1:
            axes = [axes]

        for i, category in enumerate(categories):
            matrix_data = _process_data(category, sectors, self.processes)
            ax = axes[i]
            ax.imshow(matrix_data.values, cmap="RdYlGn", vmin=-1, vmax=1)

            for j in range(len(matrix_data.index)):
                for k in range(len(matrix_data.columns)):
                    value = matrix_data.values[j, k]
                    color = (
                        "blue"
                        if value == 2
                        else "purple"
                        if value == 3
                        else "white"
                        if np.isnan(value)
                        else "red"
                        if value == -1
                        else "green"
                        if value == 1
                        else "yellow"
                    )
                    rect = plt.Rectangle(
                        (k - 0.5, j - 0.5), 1, 1, fill=True, facecolor=color, edgecolor="black", linewidth=1
                    )
                    ax.add_patch(rect)

            plt.sca(axes[i])
            ax.set_xticks(range(len(matrix_data.columns)), matrix_data.columns)

            non_nan_not_zero_rows = ((matrix_data.notna()) & (matrix_data != 0) & (matrix_data != 3)).sum(axis=1) == 1

            yticklabels = matrix_data.index.tolist()
            ax.set_yticks(range(0, len(yticklabels)))
            for label, bold in zip(ax.get_yticklabels(), non_nan_not_zero_rows):
                if bold:
                    label.set_fontweight("bold")  # Set font weight to bold for appropriate labels

            ax.set_yticks(range(len(matrix_data.index)))
            ax.set_yticklabels(yticklabels)

            ax.set_xlabel("Sectors", fontsize=12)
            ax.set_title(category.upper(), fontsize=16)

        legend_elements = [
            Patch(facecolor="white", edgecolor="black", label="No Relation"),
            Patch(facecolor="red", edgecolor="black", label="Input"),
            Patch(facecolor="green", edgecolor="black", label="Output"),
            Patch(facecolor="yellow", edgecolor="black", label="Input & Output"),
            Patch(facecolor="purple", edgecolor="black", label="Source & Input/Output"),
            Patch(facecolor="blue", edgecolor="black", label="Source Output"),
        ]
        fig.legend(handles=legend_elements, loc="upper left", fontsize=12)
        plt.tight_layout()
        plt.show()

    def get_commodity_diff(self):
        """
        This Function intends to help the user quickly identify missing
        sources or sinks in the Energy system

        input_processes allows the user to set names for processes that are
        'creating' Commodities

        output_processes allow the user to set names for processes that
        `destroy` Commodities
        """

        def add_value(original_list, adding_list):
            for item in adding_list:
                if isinstance(item, list):
                    original_list.extend(item)
                else:
                    original_list.append(item)
            return original_list

        io_dict = {"inputs": [], "outputs": []}

        for x in self.processes.values():
            inputs = x["inputs"]
            outputs = x["outputs"]

            io_dict["inputs"] = add_value(io_dict["inputs"], inputs)
            io_dict["outputs"] = add_value(io_dict["outputs"], outputs)

        # delete duplicates
        io_dict["inputs"] = np.unique(np.array(io_dict["inputs"]))
        io_dict["outputs"] = np.unique(io_dict["outputs"])

        needed_from_external_source = [x for x in io_dict["inputs"] if x not in io_dict["outputs"]]
        sink_is_necessary = [x for x in io_dict["outputs"] if x not in io_dict["inputs"]]

        return {"sink_is_necessary": sink_is_necessary, "needed_from_external_source": needed_from_external_source}
